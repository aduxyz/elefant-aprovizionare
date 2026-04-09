#!/usr/bin/env python3
"""
simulate-procurement.py
=======================
Generează  arhivă/aprovizionare-YYYY-MM-DD.xlsx  pentru o dată dată.
Folosește exclusiv elefant-erp.db — niciun fișier XLS nu este citit.

Coloane output (37): identice cu aprovizionare-sample100.xlsx.
Coloana K (Cantitate) = unități recepționate (NIR) în ziua respectivă.
Stoc Online = stoc reconstruit la data dată din stock_history.
Necesar / DOS / VZ medie/zi = calculate pe baza datelor din procurement_erp.

Utilizare:
    python simulate-procurement.py 2026-03-09

Output:
    elefant-aprovizionare-gdrive/arhivă/aprovizionare-2026-03-09.xlsx

Surse DB (elefant-erp.db):
    procurement_erp  → date produs, vânzări istorice, prețuri
    stock_history    → stoc online reconstruit la data dată
    purchases        → cantități recepționate (NIR) pe data dată
"""

import sys
import sqlite3
from datetime import datetime
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── Căi ────────────────────────────────────────────────────────────────────────
HERE    = Path(__file__).parent
ROOT    = HERE.parent
DB_FILE = ROOT / 'elefant-erp.db'
ARHIVA  = ROOT / 'data'

# ── Ordinea finală a coloanelor (identică cu sample) ──────────────────────────
OUTPUT_COLS = [
    'ElefantSKU', 'EAN', 'Cod Articol', 'Articol', 'Autor', 'Furnizor',
    'RRP', 'Reducere', 'Stoc Online', 'Necesar', 'Cantitate', 'DOS',
    'Disp. Rec.', 'SalesL1W', 'SalesLM', 'SalesL2M', 'SalesLS', 'SalesLY',
    'VZ medie/zi', 'Data Creare', 'Categorie', 'Subcategorie', 'Producator',
    'Cod Furnizor', 'Total Rez.', 'Disp. MM/Auchan', 'Colectie', 'Format',
    'Varsta', 'Cod SKU Abon', 'Achiz. All', 'Sales All', 'Epuizat',
    'Indisponibil', 'Disp. Custf.', 'Disp. Stpr.', 'Pret Achiz. Frz.',
]


# ── Funcții DB ─────────────────────────────────────────────────────────────────

def build_output_df(con: sqlite3.Connection, date_str: str) -> pd.DataFrame:
    """
    Un singur query SQL care jonează:
      - purchases (recepții pe data dată)  → Cantitate
      - procurement_erp                    → date produs
      - stock_history (cea mai recentă zi ≤ date_str) → Stoc Online reconstruit

    Returnează doar articolele recepționate în ziua respectivă (Cantitate > 0).
    """
    return pd.read_sql(f"""
        WITH
        -- Articole recepționate pe data dată și cantitățile lor
        received AS (
            SELECT article_code, SUM(quantity) AS cantitate
            FROM purchases
            WHERE SUBSTR(date, 1, 10) = '{date_str}'
              AND article_code IS NOT NULL
            GROUP BY article_code
        ),
        -- Ultima dată din stock_history ≤ date_str per articol (doar pt. articolele recepționate)
        latest_stock_date AS (
            SELECT sh.article_code, MAX(sh.stock_date) AS last_date
            FROM stock_history sh
            INNER JOIN received r ON r.article_code = sh.article_code
            WHERE sh.stock_date <= '{date_str}'
            GROUP BY sh.article_code
        ),
        -- Stocul la acea dată
        stock_at_date AS (
            SELECT sh.article_code, sh.stock_online AS stoc_la_data
            FROM stock_history sh
            INNER JOIN latest_stock_date lsd
              ON lsd.article_code = sh.article_code
             AND lsd.last_date    = sh.stock_date
        )
        SELECT
            p.sku            AS "ElefantSKU",
            p.ean            AS "EAN",
            p.article_code   AS "Cod Articol",
            p.article_desc   AS "Articol",
            p.author         AS "Autor",
            p.supplier       AS "Furnizor",
            p.rrp            AS "RRP",
            p.discount       AS "Reducere",
            COALESCE(s.stoc_la_data, p.stock_online) AS "Stoc Online",
            p.avail_rec      AS "Disp. Rec.",
            p.sales_l1w      AS "SalesL1W",
            p.sales_lm       AS "SalesLM",
            p.sales_l2m      AS "SalesL2M",
            p.sales_ls       AS "SalesLS",
            p.sales_ly       AS "SalesLY",
            p.publish_date   AS "Data Creare",
            p.category       AS "Categorie",
            p.subcategory    AS "Subcategorie",
            p.manufacturer   AS "Producator",
            p.supplier_code  AS "Cod Furnizor",
            p.total_reserved AS "Total Rez.",
            p.avail_mm_auchan AS "Disp. MM/Auchan",
            p.collection     AS "Colectie",
            p.format         AS "Format",
            p.age_group      AS "Varsta",
            p.sku_abon       AS "Cod SKU Abon",
            p.purchases_all  AS "Achiz. All",
            p.sales_all      AS "Sales All",
            p.out_of_print   AS "Epuizat",
            p.unavailable    AS "Indisponibil",
            p.avail_custf    AS "Disp. Custf.",
            p.avail_stpr     AS "Disp. Stpr.",
            p.purchase_price AS "Pret Achiz. Frz.",
            CAST(r.cantitate AS INTEGER) AS "Cantitate"
        FROM received r
        INNER JOIN procurement_erp p ON p.article_code = r.article_code
        LEFT  JOIN stock_at_date   s ON s.article_code = r.article_code
    """, con, dtype={'Cod Articol': str, 'EAN': str})


# ── Formatare Excel ────────────────────────────────────────────────────────────

def apply_formatting(path: Path, n_rows: int):
    """Header albastru+alb, Cantitate verde+bold, Data Creare format dată."""
    wb = load_workbook(path)
    ws = wb.active

    header_fill = PatternFill('solid', start_color='4472C4')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    calc_fill   = PatternFill('solid', start_color='FFFDE7')   # galben — Necesar/DOS/VZ
    cant_fill   = PatternFill('solid', start_color='D9EAD3')   # verde  — Cantitate
    cant_font   = Font(bold=True, name='Arial', size=10)
    data_font   = Font(name='Arial', size=10)

    calc_cols = {10, 12, 19}   # J=Necesar, L=DOS, S=VZ medie/zi
    cant_col  = 11             # K=Cantitate

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

    date_col = OUTPUT_COLS.index('Data Creare') + 1

    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1):
        for cell in row:
            col = cell.column
            if col == cant_col:
                cell.fill = cant_fill
                cell.font = cant_font
            elif col in calc_cols:
                cell.fill = calc_fill
                cell.font = data_font
            else:
                cell.font = data_font
            if col == date_col and cell.value:
                cell.number_format = 'DD.MM.YYYY'

    ws.freeze_panes = 'A2'
    wb.save(path)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print('Utilizare: python simulate-procurement.py YYYY-MM-DD')
        sys.exit(1)

    date_str = sys.argv[1]
    try:
        datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        print(f'Format dată invalid: {date_str}. Folosiți YYYY-MM-DD.')
        sys.exit(1)

    con = sqlite3.connect(DB_FILE)

    # ── 1. Un singur query: received × procurement_erp × stock_history ───────
    print(f'[1/3] Query principal (recepții + date produs + stoc) pentru {date_str}...')
    df_out = build_output_df(con, date_str)
    con.close()
    print(f'      {len(df_out)} articole recepționate în {date_str}')

    if df_out.empty:
        print('  Nicio recepție în această zi. Fișier gol generat.')

    # ── 2. Calcule: VZ medie/zi, Necesar, DOS ────────────────────────────────
    print('[2/3] Calculez Necesar / VZ medie/zi / DOS...')
    N = pd.to_numeric(df_out['SalesL1W'], errors='coerce').fillna(0)
    O = pd.to_numeric(df_out['SalesLM'],  errors='coerce').fillna(0)
    P = pd.to_numeric(df_out['SalesL2M'], errors='coerce').fillna(0)
    I = pd.to_numeric(df_out['Stoc Online'], errors='coerce').fillna(0)

    df_out['VZ medie/zi'] = N / 7 * 0.6 + (P - O).clip(lower=0) / 30 * 0.4
    df_out['Necesar']     = O - I
    df_out['DOS']         = np.where(df_out['VZ medie/zi'] > 0,
                                      I / df_out['VZ medie/zi'], np.nan)

    # ── 3. Sortare + export ───────────────────────────────────────────────────
    print('[3/3] Sortez și salvez...')
    sup_total = df_out.groupby('Furnizor')['Cantitate'].sum()
    df_out['_sup_total'] = df_out['Furnizor'].map(sup_total)
    df_out = df_out.sort_values(['_sup_total', 'Cantitate'], ascending=[False, False])
    df_out = df_out.drop(columns=['_sup_total'])
    df_out = df_out[OUTPUT_COLS]

    ARHIVA.mkdir(parents=True, exist_ok=True)
    out_path = ARHIVA / f'simulare-aprovizionare-{date_str}.xlsx'
    df_out.to_excel(out_path, index=False)
    apply_formatting(out_path, len(df_out))

    print(f'\n✓ Gata: {out_path}')
    print(f'  {len(df_out)} rânduri  ×  {len(OUTPUT_COLS)} coloane')


if __name__ == '__main__':
    main()
