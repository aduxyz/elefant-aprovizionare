"""
Generate procurement XLS from procurement_erp table + elefant-erp.db sales data.

Output has the same structure as EBS D.O.I. (BRUT).xlsx with 3 extra columns
inserted after Stoc online (I):
  J = Necesar   = SalesLM - StocOnline  (=O2-I2 in sheet terms)
  K = Cantitate  = computed by v4_adaptive formula
  L = DOS        = StocOnline / VZ_medie_zi  (=IFERROR(I2*30/K2,0))

Reissues (reeditări): products sharing the same title+author+supplier are grouped.
Only the most recent edition (by Data creare) is kept. Sales history from older
editions is merged into the current one.

Prerequisites: run import_ebs.py first to populate procurement_erp table.

Usage: python -m procurement.generate_orders [--date YYYY-MM-DD]
"""
import argparse
import sqlite3
import time
from datetime import datetime, timedelta
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from procurement.config import DB_PATH, TARGET_DAYS


# ── EBS D.O.I. column indices (0-based, original 33 cols A–AG) ──────
COL = {
    'ElefantSKU': 0, 'EAN': 1, 'CodArticol': 2, 'Articol': 3, 'Autor': 4,
    'Furnizor': 5, 'RRP': 6, 'Reducere': 7, 'StocOnline': 8,
    'DispReceptii': 9, 'SalesL1W': 10, 'SalesLM': 11, 'SalesL2M': 12,
    'SalesLS': 13, 'SalesLY': 14, 'DataCreare': 15, 'Categorie': 16,
    'Subcategorie': 17, 'Producator': 18, 'CodFurnizor': 19,
    'TotalRezervari': 20, 'DispMMAuchan': 21, 'Colectie': 22, 'Format': 23,
    'Varsta': 24, 'CodSKUAbon': 25, 'AchizAll': 26, 'SalesAll': 27,
    'Epuizat': 28, 'Indisponibil': 29, 'DispCustf': 30, 'DispStpr': 31,
    'PretAchizFrz': 32,
}

# Output headers: original 33 + 3 inserted after col I (pos 8)
# Original: A..I (0-8), J(9=DispReceptii)..AG(32)
# Output:   A..I (0-8), J=Necesar, K=Cantitate, L=DOS, M(=orig J)..AJ(=orig AG)
ORIGINAL_HEADERS = [
    'ElefantSKU', 'EAN', 'Cod articol', 'Articol', 'Autor', 'Furnizor',
    'RRP', 'Reducere', 'Stoc online',
    'Disponibil RECEPTII', 'SalesL1W', 'SalesLM', 'SalesL2M',
    'SalesLS', 'SalesLY', 'Data creare', 'Categorie', 'Subcategorie',
    'Producator', 'Cod furnizor', 'Total Rezervari', 'Disponibil MM Auchan',
    'Colectie', 'Format', 'Varsta', 'COD SKU ABONAMENT',
    'Achizitii All Time', 'Sales All Time', 'Epuizat', 'Indisponibil',
    'Disponibil CUSTF', 'Disponibil STPR', 'Pret achiz frz',
]

INSERTED_HEADERS = ['Necesar', 'Cantitate', 'DOS']
INSERT_POS = 9  # insert after col I (index 8), before original col J


# ── FORMULA (v4_adaptive, hardcoded after backtesting) ───────────────

def compute_daily_rate(L1W, LM, L2M, LS, LY):
    w_rate = L1W / 7.0 if L1W > 0 else 0
    m_rate = LM / 30.0 if LM > 0 else 0
    prev_m_rate = max(0, (L2M - LM)) / 30.0
    s_rate = LS / 180.0 if LS > 0 else 0
    y_rate = LY / 365.0 if LY > 0 else 0

    # Weekly post-spike: luna actuală mare, săptămâna curentă a revenit la normal
    if LM > 50 and w_rate < m_rate * 0.1 and m_rate > prev_m_rate * 3:
        return w_rate
    # Monthly post-spike: luna anterioară a explodat (5x+), acum a revenit
    if m_rate > 0 and prev_m_rate > m_rate * 5:
        return m_rate * 1.1
    if L1W == 0 and LM == 0:
        return y_rate * 0.15 if LY > 10 else 0

    momentum = (w_rate / m_rate) if m_rate > 0 else (2.0 if w_rate > 0 else 0)

    if momentum > 1.5:
        daily = w_rate * 0.6 + m_rate * 0.3 + s_rate * 0.1
    elif momentum < 0.5 and LY > 20:
        daily = m_rate * 0.4 + s_rate * 0.3 + y_rate * 0.3
    else:
        daily = w_rate * 0.3 + m_rate * 0.35 + s_rate * 0.2 + y_rate * 0.15

    return daily * 1.1


# ── REISSUE HANDLING ─────────────────────────────────────────────────

def build_reissue_groups(rows):
    """
    Group rows by (title, author, supplier) normalized.
    Returns dict: group_key -> list of row indices, sorted by DataCreare desc.
    """
    groups = defaultdict(list)
    for idx, row in enumerate(rows):
        title = str(row[COL['Articol']] or '').strip().lower()
        author = str(row[COL['Autor']] or '').strip().lower()
        supplier = str(row[COL['Furnizor']] or '').strip().lower()
        if not title:
            continue
        key = (title, author, supplier)
        groups[key].append(idx)

    # Sort each group by DataCreare descending (most recent first)
    for key in groups:
        groups[key].sort(
            key=lambda i: _parse_date(rows[i][COL['DataCreare']]),
            reverse=True
        )

    return groups


def _parse_date(val):
    if val is None:
        return datetime.min
    if isinstance(val, datetime):
        return val
    try:
        return datetime.strptime(str(val)[:10], '%Y-%m-%d')
    except (ValueError, TypeError):
        return datetime.min


def merge_reissue_sales(rows, groups):
    """
    For each reissue group, merge sales columns from older editions into
    the most recent one. Mark older editions for exclusion.

    Returns:
        keep_indices: set of row indices to keep (most recent per group)
        merged_sales: dict idx -> merged sales values for kept rows
    """
    keep_indices = set()
    merged_sales = {}

    sales_cols = ['SalesL1W', 'SalesLM', 'SalesL2M', 'SalesLS', 'SalesLY',
                  'AchizAll', 'SalesAll']

    for key, indices in groups.items():
        if len(indices) == 1:
            keep_indices.add(indices[0])
            continue

        # Keep the most recent edition (first after sort)
        current_idx = indices[0]
        keep_indices.add(current_idx)

        # Sum sales from all editions
        totals = {}
        for col_name in sales_cols:
            total = 0
            for idx in indices:
                val = rows[idx][COL[col_name]]
                total += float(val) if val else 0
            totals[col_name] = total

        merged_sales[current_idx] = totals

    # Add rows that aren't in any group (shouldn't happen if all have titles)
    all_grouped = set()
    for indices in groups.values():
        all_grouped.update(indices)
    for idx in range(len(rows)):
        if idx not in all_grouped:
            keep_indices.add(idx)

    return keep_indices, merged_sales


# ── MAIN ─────────────────────────────────────────────────────────────

def generate(db_path=DB_PATH, ref_date=None):
    if ref_date is None:
        ref_date = datetime.now().strftime('%Y-%m-%d')

    t0 = time.time()

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    # 1. Load EBS data from procurement_erp table
    print("Loading procurement_erp from DB...")
    db_rows = conn.execute("""
        SELECT sku, ean, article_code, article_desc, author, supplier, rrp, discount,
               stock_online, avail_rec, sales_l1w, sales_lm, sales_l2m, sales_ls, sales_ly,
               publish_date, category, subcategory, manufacturer, supplier_code,
               total_reserved, avail_mm_auchan, collection, format, age_group, sku_abon,
               purchases_all, sales_all, out_of_print, unavailable, avail_custf, avail_stpr,
               purchase_price
        FROM procurement_erp
    """).fetchall()
    all_rows = [list(r) for r in db_rows]
    print(f"  {len(all_rows)} rows loaded ({time.time()-t0:.1f}s)")

    # 2. Handle reissues
    print("Detecting reissues...")
    groups = build_reissue_groups(all_rows)
    multi_groups = {k: v for k, v in groups.items() if len(v) > 1}
    print(f"  {len(multi_groups)} reissue groups ({sum(len(v) for v in multi_groups.values())} rows)")

    keep_indices, merged_sales = merge_reissue_sales(all_rows, groups)
    print(f"  Keeping {len(keep_indices)} rows (removed {len(all_rows) - len(keep_indices)} older editions)")

    # 3. Load sales windows from DB for computing Cantitate
    print(f"Loading sales data from DB (ref_date={ref_date})...")
    ref_dt = datetime.strptime(ref_date, '%Y-%m-%d')

    db_windows = {}
    for name, days in [('L1W', 7), ('LM', 30), ('L2M', 60), ('LS', 180), ('LY', 365)]:
        start = (ref_dt - timedelta(days=days)).strftime('%Y-%m-%d')
        rows_db = conn.execute(
            "SELECT article_code, SUM(qty) as total FROM daily_sales "
            "WHERE date >= ? AND date < ? GROUP BY article_code",
            (start, ref_date)
        ).fetchall()
        for r in rows_db:
            ac = r['article_code']
            if ac not in db_windows:
                db_windows[ac] = {'L1W': 0, 'LM': 0, 'L2M': 0, 'LS': 0, 'LY': 0}
            db_windows[ac][name] = r['total']
    print(f"  {len(db_windows)} products with sales history")

    # 4. Build output rows
    print("Computing Necesar/Cantitate/DOS...")
    output_rows = []
    for idx in sorted(keep_indices):
        row = list(all_rows[idx])

        # Apply merged sales if this is a reissue with merged data
        if idx in merged_sales:
            for col_name, val in merged_sales[idx].items():
                row[COL[col_name]] = val

        # Read values for formulas
        stoc_online = float(row[COL['StocOnline']] or 0)
        disp_rec = float(row[COL['DispReceptii']] or 0)
        article_code = str(row[COL['CodArticol']] or '')

        # Use DB windows (computed from daily_sales up to ref_date) — always
        # more recent than EBS snapshot. Falls back to EBS if no DB history.
        if article_code in db_windows:
            w = db_windows[article_code]
            L1W = w['L1W']
            LM  = w['LM']
            L2M = w['L2M']
            LS  = w['LS']
            LY  = w['LY']
        else:
            L1W = float(row[COL['SalesL1W']] or 0)
            LM  = float(row[COL['SalesLM']]  or 0)
            L2M = float(row[COL['SalesL2M']] or 0)
            LS  = float(row[COL['SalesLS']]  or 0)
            LY  = float(row[COL['SalesLY']]  or 0)

        # J = Necesar = SalesLM - StocOnline (=O2-I2 in output)
        necesar = LM - stoc_online

        # S = VZ medie/zi (avg daily sales)
        vz_medie = compute_daily_rate(L1W, LM, L2M, LS, LY)

        # K = Cantitate = formula result - stoc - receptii
        # If already overstocked (Necesar < 0), don't order regardless of historical demand
        if necesar < 0:
            cantitate = 0
        else:
            cantitate = max(0, round(vz_medie * TARGET_DAYS - stoc_online - disp_rec))

        # L = DOS = StocOnline / VZ_medie_zi (=IFERROR(I2/S2,0))
        dos = round(stoc_online / vz_medie, 1) if vz_medie > 0 else 0

        # Build output: original A-I, then J/K/L, then original J-AG
        out = row[:INSERT_POS] + [necesar, cantitate, dos] + row[INSERT_POS:]
        output_rows.append(out)

    # Sort: primary = supplier desc by total Cantitate, secondary = Cantitate desc
    K = INSERT_POS + 1  # Cantitate index in output row (0-based)
    F = COL['Furnizor']  # Furnizor index (0-based, same position in output row)

    supplier_totals = defaultdict(int)
    for r in output_rows:
        supplier_totals[r[F]] += r[K] if isinstance(r[K], (int, float)) else 0

    output_rows.sort(key=lambda r: (
        -supplier_totals[r[F]],
        -(r[K] if isinstance(r[K], (int, float)) else 0),
    ))

    # 5. Write output XLS
    output_path = f"aprovizionare-{ref_date}.xlsx"
    print(f"Writing {len(output_rows)} rows to {output_path}...")

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'aprovizionare'

    # Build full header list before reordering
    out_headers_full = ORIGINAL_HEADERS[:INSERT_POS] + INSERTED_HEADERS + ORIGINAL_HEADERS[INSERT_POS:]
    # I:R = indices 8..17 (0-based) moved after C (index 2)
    # New order: [0,1,2] + [8..17] + [3..7] + [18..end]
    n = len(out_headers_full)
    col_order = list(range(3)) + list(range(8, 18)) + list(range(3, 8)) + list(range(18, n))
    out_headers = [out_headers_full[i] for i in col_order]

    # Column positions in reordered output (1-based)
    # old index 8=StocOnline → new pos 3 → col 4
    # old index 9=Necesar    → new pos 4 → col 5
    # old index 10=Cantitate → new pos 5 → col 6
    # old index 11=DOS       → new pos 6 → col 7
    # old index 14=SalesLM   → new pos 9 → col 10
    # old index 6=RRP        → new pos 17 → col 18
    # old index 35=PretAchiz → still last → col n
    def new_col1(old_0based):
        return col_order.index(old_0based) + 1

    COL1_STOC    = new_col1(8)   # StocOnline (I in formula =O-I)
    COL1_NECESAR = new_col1(9)   # Necesar
    COL1_CANT    = new_col1(10)  # Cantitate
    COL1_DOS     = new_col1(11)  # DOS
    COL1_SALESLM = new_col1(14)  # SalesLM (O in formula =O-I)
    COL1_RRP        = new_col1(6)
    COL1_PRET_ACHIZ = new_col1(35)

    ws_out.append(out_headers)
    for col in range(1, len(out_headers) + 1):
        ws_out.cell(row=1, column=col).font = Font(bold=True)

    # Freeze first column
    ws_out.freeze_panes = 'A2'

    # Write formulas (not static values) for Necesar/DOS cells
    for r_idx, out_row in enumerate(output_rows, 2):
        reordered = [out_row[i] for i in col_order]
        for c_idx, val in enumerate(reordered, 1):
            ws_out.cell(row=r_idx, column=c_idx, value=val)

        # Number formats
        ws_out.cell(row=r_idx, column=COL1_RRP).number_format        = '0.00'
        ws_out.cell(row=r_idx, column=COL1_PRET_ACHIZ).number_format = '0.00'
        ws_out.cell(row=r_idx, column=COL1_DOS).number_format        = '0.0'

        stoc_letter = get_column_letter(COL1_STOC)
        cant_letter = get_column_letter(COL1_CANT)
        lm_letter   = get_column_letter(COL1_SALESLM)

        # Necesar = SalesLM - StocOnline
        ws_out.cell(row=r_idx, column=COL1_NECESAR,
                    value=f'={lm_letter}{r_idx}-{stoc_letter}{r_idx}')
        # DOS = IFERROR(StocOnline*30/Cantitate, 0)
        ws_out.cell(row=r_idx, column=COL1_DOS,
                    value=f'=IFERROR({stoc_letter}{r_idx}*30/{cant_letter}{r_idx},0)')

    # Auto-width for first 15 cols
    for col_idx in range(1, min(len(out_headers) + 1, 37)):
        letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws_out.cell(row=r, column=col_idx).value or ''))
            for r in range(1, min(52, len(output_rows) + 2))
        )
        ws_out.column_dimensions[letter].width = min(max_len + 2, 40)

    wb_out.save(output_path)

    conn.close()

    elapsed = time.time() - t0
    total_cant = sum(1 for r in output_rows if r[INSERT_POS + 1] > 0)
    total_qty = sum(r[INSERT_POS + 1] for r in output_rows if isinstance(r[INSERT_POS + 1], (int, float)))
    print(f"\nDone in {elapsed:.1f}s")
    print(f"  Total rows: {len(output_rows)}")
    print(f"  Products with Cantitate > 0: {total_cant}")
    print(f"  Total Cantitate: {total_qty:,.0f}")
    print(f"  Reissue groups merged: {len(multi_groups)}")
    print(f"  Output: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='Generate procurement orders')
    parser.add_argument('--date', default=None, help='Reference date (YYYY-MM-DD), default=today')
    parser.add_argument('--db', default=DB_PATH, help='Path to elefant-erp.db')
    args = parser.parse_args()
    generate(db_path=args.db, ref_date=args.date)


if __name__ == '__main__':
    main()
